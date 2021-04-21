import openpyxl as xl


#Open source file for Ski Sales
fileName1 = "C:\\Users\\tcreps\\Desktop\\AHK\\Report Scripts\\Jon Sale.xlsx"
wb1 = xl.load_workbook(fileName1)
ws1 = wb1.worksheets[0]

#open desitination file
fileName2 = "H:\\Jon Tyler\\Ski Sales2020.xlsx"
wb2 = xl.load_workbook(fileName2)
ws2 = wb2.worksheets[1]


# Define max rows and columns needed to write
mr = ws1.max_row
mc = 8

# Copy data from columns A : G from wb1 to wb2
for i in range (1, mr + 1):
    for j in range (1, mc + 1):
        c = ws1.cell(row = i, column = j)

        ws2.cell(row = i, column = j).value = c.value


# Save output as new Document
wb2.save('H:\\Jon Tyler\\Ski Sales2020.xlsx')

