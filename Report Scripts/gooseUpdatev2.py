import openpyxl as xl
from datetime import datetime


gooseFile1 = "C:\\Users\\tcreps\\Desktop\\AHK\\Report Scripts\\Jon Sale.xlsx"
wb1 = xl.load_workbook(gooseFile1)
ws1 = wb1.worksheets[0]

gooseFile2 = "H:\\Jon Tyler\\Canada Goose sales.xlsx"
wb2 = xl.load_workbook(gooseFile2)
ws2 = wb2.worksheets[5]

now = datetime.now()

# Define max rows and columns needed to write
mr = ws1.max_row
mc = 7

# Copy data from columns A : G from wb1 to wb2
for i in range (1, mr + 1):
    for j in range (1, mc + 1):
        c = ws1.cell(row = i, column = j)

        ws2.cell(row = i, column = j).value = c.value
        ws2.cell(row = i, column = j + 1).value = now.strftime("%B")

wb2.save('H:\\Jon Tyler\\Canada Goose sales.xlsx')

# wb2.save('C:\\Users\\tcreps\\Desktop\\Canada Goose sales.xlsx')




