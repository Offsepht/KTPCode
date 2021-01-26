import pandas as pd
from openpyxl import workbook
from openpyxl import load_workbook


df = pd.read_csv(element, "|")
df = df.drop([0,0])

wb = load_workbook("C:\\Users\\tcreps\\Desktop\\Ski Sales2020")
element = load_workbook('C:\\Users\\tcreps\\Desktop\\Jon Sales.xslx')
ws = wb.worksheet[1]
ws2 = element.worksheet[0]

ws_tables = []
for table in ws._tables:
    ws_tables.append(tables)
    print(table.name, table.ref)






df.to_excel('C:\\Users\\tcreps\\Desktop\\Ski Sales2020', 'Jan21', index=0)