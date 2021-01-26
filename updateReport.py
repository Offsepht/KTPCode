import openpyxl as xl

fileName1 = "C:\\Users\\tcreps\\Desktop\\Jon Sale.xlsx"
wb1 = xl.load_workbook(fileName1)
ws1 = wb1.worksheets[0]


fileName2 = "C:\\Users\\tcreps\\Desktop\\Ski Sales2020.xlsx"
wb2 = xl.load_workbook(fileName2)
ws2 = wb2.worksheets[1]


mr = ws1.max_row
mc = 6

for i in range (1, mr + 1):
    for j in range (1, mc + 1):
        c = ws1.cell(row = i, column = j)

        ws2.cell(row = i, column = j).value = c.value

wb2.save('C:\\Users\\tcreps\\Desktop\\Ski Sales update.xlsx')

